import csv
import os
import pdb

pdb.set_trace()
import pandas as pd
import openpyxl
import datetime
import re
from pymongo import MongoClient

# MongoDB connection URI
mongo_uri = "mongodb://localhost:27017/"
mongo_dbname = "yoanonedatabase"

# Create a MongoDB client and connect to the database
mongo_client = MongoClient(mongo_uri)
mongo_db = mongo_client[mongo_dbname]


def load_conditions_from_excel(file_path, column_number):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    conditions = []
    for row_idx, row in enumerate(sheet.iter_rows(min_col=column_number, max_col=column_number, values_only=True),
                                  start=1):
        if row_idx == 1:  # Skip the first row (header)
            continue
        for cell_value in row:
            if cell_value is not None:
                conditions.append(cell_value)
            else:
                break
    return conditions


def fetch_all_data_mongodb(collection):
    try:
        # Fetch all data from the MongoDB collection
        data = list(collection.find())
        print("Data fetched")
        return data
    except Exception as e:
        print(f"Error fetching data from MongoDB: {e}")
        return []


def fetch_data_with_conditions(collection, third_conditions,
                               mapped_fourth_conditions, seven_conditions,
                               email_conditions, jt_link_conditions,
                               fl_domain_conditions, fl_company_conditions):
    try:
        # Construct the query based on conditions
        query = {}

        # Condition 5: Exclude rows where 'Email' is in email_conditions
        if email_conditions:
            query['Email'] = {"$nin": email_conditions}

        # Condition 6: Exclude rows where 'Job_Title_Link' is in jt_link_conditions
        if jt_link_conditions:
            query['Job_Title_Link'] = {"$nin": jt_link_conditions}

        # Condition 7: Exclude rows where 'FirstLastDomain' is in fl_domain_conditions
        if fl_domain_conditions:
            query['FirstLastDomain'] = {"$nin": fl_domain_conditions}

        # Condition 8: Exclude rows where 'FirstLastCompany' is in fl_company_conditions
        if fl_company_conditions:
            query['FirstLastCompany'] = {"$nin": fl_company_conditions}

        # Condition 2: Filter by 'COUNTRY' using $in
        if third_conditions:
            query['COUNTRY'] = {"$in": third_conditions}

        # Condition 3: Filter by 'Employee_Size' using $in
        if mapped_fourth_conditions:
            query['Employee_Size'] = {"$in": mapped_fourth_conditions}

        # Condition 4: Exclude rows where 'domain' is in seven_conditions
        if seven_conditions:
            query['domain'] = {"$nin": seven_conditions}

        # Fetch data from MongoDB collection with the specified conditions
        data = list(collection.find(query))
        print("Data fetched")
        return data
    except Exception as e:
        print(f"Error fetching data from MongoDB: {e}")
        return []


def main():
    try:
        file_path = r'\\yoandc\Campaigns\#Yoanone\dipesh\JT_For_Extraction.xlsm'
        # Load input conditions from the first sheet of Excel
        workbook = openpyxl.load_workbook(file_path)
        print("Execution Start")
        output_file_name = input("Enter file name with extension:")
        output_file_path = fr'\yoandc\Campaigns\#Yoanone\dipesh\{output_file_name}'

        # Load input conditions from Excel
        first_conditions = load_conditions_from_excel(file_path, column_number=1)
        second_conditions = load_conditions_from_excel(file_path, column_number=2)
        third_conditions = load_conditions_from_excel(file_path, column_number=3)
        fourth_conditions = load_conditions_from_excel(file_path, column_number=4)
        fifth_conditions = load_conditions_from_excel(file_path, column_number=5)
        sixth_conditions = load_conditions_from_excel(file_path, column_number=6)
        seven_conditions = load_conditions_from_excel(file_path, column_number=7)
        email_conditions = load_conditions_from_excel(file_path, column_number=8)
        jt_link_conditions = load_conditions_from_excel(file_path, column_number=9)
        fl_domain_conditions = load_conditions_from_excel(file_path, column_number=10)
        fl_company_conditions = load_conditions_from_excel(file_path, column_number=11)

        # Load 'Job Level' sheet from Excel
        job_level_sheet = workbook['Job Level']
        mapped_conditions = []
        for condition in first_conditions:
            for row in job_level_sheet.iter_rows(values_only=True):
                if condition.lower() in str(row[0]).lower():
                    mapped_conditions.append(row[1])

        print(mapped_conditions)
        print(third_conditions)

        company_size = workbook['companySize']
        mapped_fourth_conditions = []
        for condition in fourth_conditions:
            for row in company_size.iter_rows(values_only=True):
                if condition.lower() in str(row[0]).lower():
                    mapped_fourth_conditions.append(row[1])

        second_sheet = workbook['JT']  # Provide the actual name of your second sheet
        matched_column_indexes = []
        matched_conditions = []
        # Iterate through rows in the second sheet
        for row_idx, row in enumerate(second_sheet.iter_rows(min_row=1, max_row=1, values_only=True), start=1):
            # Iterate through cells in the row
            for idx, cell_value in enumerate(row, start=1):
                # Check if the cell value matches any of the second_conditions
                if cell_value in second_conditions:
                    matched_column_indexes.append(idx)

        # Iterate through matched columns
        for matched_column_index in matched_column_indexes:
            # Get all conditions from the current matched column
            for row_idx, cell_value in enumerate(
                    second_sheet.iter_rows(min_col=matched_column_index, max_col=matched_column_index,
                                           values_only=True),
                    start=1):
                if row_idx == 1:  # Skip the first row (header)
                    continue
                if cell_value[0] is not None:
                    matched_conditions.append(cell_value[0])
                else:
                    break

        start_time = datetime.datetime.now().strftime("%M%S")
        start_time_int = int(start_time)
        print("Start time:", start_time)
        all_results_df = pd.DataFrame()  # Create an empty DataFrame

        count = 1
        for i in range(1, 22):
            # Process data in batches
            collection_name = f"yoan_one_{count}"
            print(collection_name)

            # Fetch all data from MongoDB collection
            # all_result = fetch_all_data_mongodb(mongo_db[collection_name])
            all_result = fetch_data_with_conditions(
                mongo_db[collection_name],
                third_conditions,
                mapped_fourth_conditions, seven_conditions,
                email_conditions, jt_link_conditions,
                fl_domain_conditions, fl_company_conditions
            )

            count += 1
            # Create a DataFrame from the results
            header = ['Date', 'Salutation', 'First_Name', 'Last_Name', 'Email', 'Company_Name', 'Address_1',
                      'City', 'State', 'Zip_Code', 'COUNTRY', 'Industry', 'Standard_Industry',
                      'Job_Title', 'Job_Title_Level', 'Job_Title_Department', 'Employee_Size', 'Revenue_Size',
                      'Phone_NO', 'Direct_Dial_Extension', 'SIC_Code', 'NAICS_Code', 'Job_Title_Link',
                      'Employee_Size_Link',
                      'Revenue_Size_Link', 'VV_Status', 'Final_Status', 'id', 'domain', 'FirstLastDomain',
                      'FirstLastCompany']

            all_results_df = pd.DataFrame(all_result, columns=header)
            # Concatenate the new results with the existing DataFrame
            all_results_df = all_results_df.drop_duplicates('Job_Title_Link')
            all_results_df = all_results_df.drop_duplicates('Email')
            all_results_df = all_results_df.drop_duplicates('FirstLastDomain')
            all_results_df = all_results_df.drop_duplicates('FirstLastCompany')
            try:
                print("Data filtration start")

                # Function to apply regex pattern using re.findall
                def apply_regex(column, pattern):
                    return column.apply(lambda x: bool(re.findall(pattern, str(x), flags=re.IGNORECASE)))

                try:
                    if sixth_conditions:
                        all_results_df = all_results_df[all_results_df['domain'].isin(sixth_conditions)]
                        print("six", all_results_df.shape)
                except Exception as e:
                    print("Exception in Tal", e)

                # try:
                #     if third_conditions:
                #         all_results_df = all_results_df[all_results_df['COUNTRY'].isin(third_conditions)]
                #         print("third", all_results_df.shape)
                # except Exception as e:
                #     print("Exception in country", e)
                #
                # try:
                #     if mapped_fourth_conditions:
                #         all_results_df = all_results_df[all_results_df['Employee_Size'].isin(mapped_fourth_conditions)]
                # except Exception as e:
                #     print("Exception in employee size", e)
                #
                # try:
                #     # Check if seventh_conditions is not empty, apply it to the DataFrame
                #     if seven_conditions:
                #         all_results_df = all_results_df[~all_results_df['domain'].isin(seven_conditions)]
                # except Exception as e:
                #     print("Exception in suppression", e)
                # try:
                #     # Apply the fifth condition using str.contains
                #     if mapped_conditions:
                #         condition_series = []
                #         for condition in mapped_conditions:
                #             pattern = f".*{condition}.*"
                #             condition_series.append(
                #                 all_results_df['Job_Title'].str.contains(pattern, case=False, na=False, regex=True))
                #         if condition_series:
                #             # Combine conditions using logical OR
                #             final_condition = pd.DataFrame(condition_series).any(axis=0)
                #             all_results_df = all_results_df[final_condition]
                # except Exception as e:
                #     print("Error in industry condition:", e)
                #
                # try:
                #     # Apply the fifth condition using str.contains
                #     if matched_conditions:
                #         condition_series = []
                #         for condition in matched_conditions:
                #             pattern = f".*{condition}.*"
                #             condition_series.append(
                #                 all_results_df['Job_Title'].str.contains(pattern, case=False, na=False, regex=True))
                #         if condition_series:
                #             # Combine conditions using logical OR
                #             final_condition = pd.DataFrame(condition_series).any(axis=0)
                #             all_results_df = all_results_df[final_condition]
                # except Exception as e:
                #     print("Error in industry condition:", e)
                # try:
                #     # Apply the condition using str.contains
                #     if mapped_conditions and matched_conditions:
                #         condition_series = []
                #
                #         for condition in mapped_conditions:
                #             pattern = f".*{condition}.*"
                #
                #             for condition1 in matched_conditions:
                #                 pattern1 = f".*{condition1}.*"  # Use pattern1 for the second condition
                #
                #                 # Use logical AND to check both patterns with any characters before and after
                #                 combined_pattern = f".*{pattern}.*&.*{pattern1}.*"
                #                 condition_series.append(
                #                     all_results_df['Job_Title'].str.contains(combined_pattern, case=False, na=False,
                #                                                              regex=True)
                #                 )
                #
                #         if condition_series:
                #             # Combine conditions using logical OR
                #             final_condition = pd.DataFrame(condition_series).any(axis=0)
                #             all_results_df = all_results_df[final_condition]
                #
                # except Exception as e:
                #     print("Error in industry condition:", e)
                condition_series = []

                def match(matched_conditions, mapped_conditions, a):
                    for i in mapped_conditions:
                        combined_pattern = f"{matched_conditions[a]}&{i}"
                        condition_series.append(
                            all_results_df['Job_Title'].str.contains(combined_pattern, case=False, na=False,
                                                                     regex=True)
                        )
                    if len(matched_conditions) > a:
                        a = a + 1
                        match(matched_conditions, mapped_conditions, a)
                    else:
                        return


                try:
                    # Apply the fifth condition using re.findall
                    if fifth_conditions:
                        for condition in fifth_conditions:
                            pattern = f"{condition}.*"
                            all_results_df = all_results_df[~apply_regex(all_results_df['Industry'], pattern)]
                except Exception as e:
                    print("Exception in Industry", e)

                print(all_results_df.shape)
                # try:
                #     # Check if email_conditions is not empty, apply it to the DataFrame
                #     if email_conditions:
                #         all_results_df = all_results_df[
                #             ~all_results_df['Email'].isin(email_conditions)]
                # except Exception as e:
                #     print("Exception in email suppression", e)
                #
                # try:
                #     # Check if jt_link_conditions is not empty, apply it to the DataFrame
                #     if jt_link_conditions:
                #         all_results_df = all_results_df[~all_results_df['Job_Title_Link'].isin(jt_link_conditions)]
                # except Exception as e:
                #     print("Exception in JT link suppression", e)
                #
                # try:
                #     # Check if fl_domain_conditions is not empty, apply it to the DataFrame
                #     if fl_domain_conditions:
                #         all_results_df = all_results_df[
                #             ~all_results_df['FirstLastDomain'].isin(fl_domain_conditions)]
                # except Exception as e:
                #     print("Exception in FL_domain suppression", e)
                #
                # try:
                #     # Check if fl_company_conditions is not empty, apply it to the DataFrame
                #     if fl_company_conditions:
                #         all_results_df = all_results_df[
                #             ~all_results_df['FirstLastCompany'].isin(fl_company_conditions)]
                # except Exception as e:
                #     print("Exception in FL_Company suppression", e)
                print(output_file_path)
                try:
                    # with open(output_file_path, 'a') as file:
                    #     # Add our players info in csv file which already created
                    #     writer = csv.writer(file)
                    #     writer.writerow(output_file_path)
                    length = pd.read_excel(output_file_path)
                    all_results_df.to_excel(output_file_path, mode='a', startrow=len(length))
                except Exception as e:
                    print(e)
                all_results_df = pd.DataFrame()  # Create an empty DataFrame
            except Exception as e:
                print(e)

        end_time = datetime.datetime.now().strftime("%M%S")
        end_time_int = int(end_time)
        print("end time:", end_time)
        total_time_script_takes = abs(start_time_int - end_time_int)
        print("Total time takes:", total_time_script_takes)
        # Print message
        print(f"Data saved into Excel file: {output_file_name}")

    except Exception as e:
        print(e)

    finally:
        # Close MongoDB connection
        mongo_client.close()


if __name__ == '__main__':
    main()
